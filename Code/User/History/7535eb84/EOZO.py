import os
import boto3
from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
from datetime import datetime
from datetime import timedelta
from openpyxl.drawing.image import Image

cell_tracking = 1

wb = Workbook()
ws = wb.active

# Set up AWS credentials
session = boto3.Session(profile_name='mcdonalds')
cloudwatch = session.client('cloudwatch')

today = (datetime.now()).strftime("%d/%m/%Y %H:%M:%S")
week_ago = datetime.now() - timedelta(days=7)

ec2_unit_list = ['CPU Utilization (%)', 'Gigabytes', 'Gigabytes', 'Count']
ec2_metric_list = ['CPUUtilization', 'NetworkIn', 'NetworkOut', 'StatusCheckFailed_System']
ec2_dict = {
    'mcclient1': 'i-0fd9a6412bcd42050',
    'mcclient3': 'i-0d61bd2a498aa824f',
    'mcclient5': 'i-014c748b8e83ff439'
}

rds_unit_list = ['CPU Utilization (%)', 'Count', 'Count/Second', 'Count/Second']
rds_metric_list = ['CPUUtilization', 'DatabaseConnections', 'ReadIOPS', 'WriteIOPS']
rds_dict = {
    'mcdonalds-live-reader1': 'mcdonalds-live-reader1',
    'mcdonalds-live-reader2': 'mcdonalds-live-reader2',
    'mcdonalds-live-writer': 'mcdonalds-live-writer'
}

redis_unit_list = ['CPU Utilization (%)', 'Count', 'Count', 'Count']
redis_metric_list = ['EngineCPUUtilization', 'CurrConnections', 'CurrItems', 'CacheHits']
redis_dict = {
    'mclive-redis-001': 'mclive-redis-001',
    'mclive-redis-002': 'mclive-redis-002'
}

print(cell_tracking//20 + 2)

def get_cloudwatch_metric_graphs(namespace, metric_name, identifier, instance_id, unit):
    global cell_tracking
    response = cloudwatch.get_metric_statistics(
        Namespace=namespace,
        MetricName=metric_name,
        Dimensions=[{
            'Name': identifier,
            'Value': instance_id
        }],
        StartTime=week_ago - timedelta(days=7),
        EndTime=week_ago,
        Period=1200,
        Statistics=['Average']
    )
    response['Datapoints'].sort(key=lambda x: x['Timestamp'])
    values = [point['Average'] for point in response['Datapoints']]
    value_averages_last_week = sum(values) / len(values)
    response = cloudwatch.get_metric_statistics(
        Namespace=namespace,
        MetricName=metric_name,
        Dimensions=[{
            'Name': identifier,
            'Value': instance_id
        }],
        StartTime=week_ago,
        EndTime=today,
        Period=1200,
        Statistics=['Average']
    )

    response['Datapoints'].sort(key=lambda x: x['Timestamp'])
    values = [point['Average'] for point in response['Datapoints']]
    value_averages = sum(values) / len(values)
    if metric_name == "NetworkIn" or metric_name == "NetworkOut":
        value_averages = value_averages / 1024 / 1024
        value_averages_last_week = value_averages_last_week / 1024 / 1024

    # Prepare data
    time_stamps = [point['Timestamp'] for point in response['Datapoints']]

    # Create a plot
    plt.figure(figsize=(14, 4))
    plt.plot_date(time_stamps, values, '-')
    print(time_stamps, values)

    # Format the plot
    plt.title('AWS CloudWatch Metric - ' + metric_name)
    plt.xlabel('Time')
    plt.ylabel(unit)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.gca().xaxis.set_major_formatter(DateFormatter('%Y-%m-%d'))
    plt.gcf().autofmt_xdate()  # Rotation

    # Save the plot as an image
    plt.savefig(instance_id + metric_name + ".png")
    plt.close()
    img = Image(instance_id + metric_name + '.png')

    ws.add_image(img, 'A' + str(cell_tracking))  # Adjust the cell location as needed
    ws_pointer = cell_tracking // 20 + 2
    ws["X" + str(ws_pointer)] = metric_name
    ws["Y" + str(ws_pointer)] = value_averages
    ws["Z" + str(ws_pointer)] = value_averages_last_week

    cell_tracking = cell_tracking + 20
def get_metrics(namespace, dict, metric_list, unit_list, indentifier):
    global ws
    global cell_tracking
    for instance in dict.values():
        for key, value in dict.items():
            if value == instance:
                ws = wb.create_sheet(title=key)
                ws["Y1"] = "Last Week"
                ws["Z1"] = "This Week"
        for i, metric in enumerate(metric_list):
            get_cloudwatch_metric_graphs(namespace, metric, indentifier, instance, unit_list[i])
        ws["AA1"] = 'Anomaly'
        ws["AA2"] = '=Z2-Y2'
        ws["AA3"] = '=Z3-Y3'
        ws["AA4"] = '=Z4-Y4'
        ws["AA5"] = '=Z5-Y5'
        cell_tracking = 1

get_metrics('AWS/EC2', ec2_dict, ec2_metric_list, ec2_unit_list, 'InstanceId')
get_metrics('AWS/RDS', rds_dict, rds_metric_list, rds_unit_list, 'DBInstanceIdentifier')
get_metrics('AWS/ElastiCache', redis_dict, redis_metric_list, redis_unit_list, 'CacheClusterId')

today = (datetime.now()).strftime("%d-%m-%Y")
week_ago = week_ago.strftime("%d-%m-%Y")

wb.remove(wb["Sheet"])
wb.save(os.path.join(script_dir, "mcdonalds_anomaly_" + week_ago + "-" + today + ".xlsx"))

all_items = os.listdir(script_dir)
png_files = [item for item in all_items if item.endswith('.png')]
for file_name in png_files:
    file_path = os.path.join(script_dir, file_name)
    try:
        os.remove(file_path)
        print(f"Deleted: {file_path}")
    except Exception as e:
        print(f"Error deleting {file_path}: {e}")
