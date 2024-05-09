import os
import boto3
from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
from datetime import datetime
from datetime import timedelta
from openpyxl.drawing.image import Image

cell_tracking = 1

wb = Workbook()
ws = wb.active

# Assume the role of the relevant AWS account
sts_client = boto3.client('sts')
assumed_role = sts_client.assume_role(
    RoleArn="arn:aws:iam::146976730351:role/OrganizationAccountAccessRole",
    RoleSessionName="GetCloudwatchMetrics"
)
credentials = assumed_role['Credentials']

# Set up AWS credentials
session = boto3.Session(
    aws_access_key_id=credentials['AccessKeyId'],
    aws_secret_access_key=credentials['SecretAccessKey'],
    aws_session_token=credentials['SessionToken']
)

# Set up the cloudwatch client for this session
cloudwatch = session.client('cloudwatch')

script_dir = os.path.dirname(os.path.abspath(__file__))


today = (datetime.now()).strftime("%d/%m/%Y %H:%M:%S")
week_ago = datetime.now() - timedelta(days=7)

ec2_unit_list = ['CPU Utilization (%)', 'Gigabytes', 'Gigabytes', 'Count']
ec2_metric_list = ['CPUUtilization', 'NetworkIn', 'NetworkOut', 'StatusCheckFailed_System']
ec2_dict = {
    'GrowLiveClient1': 'i-0d7b4f48c53321810',
    'GrowLiveClient3': 'i-0d1ae698ae0efb226',
    'GrowLiveClient5': 'i-02e0588ed5e9b0732',
    'GrowLiveClient14': 'i-01dc536f769291c3d',
    'GrowLiveClient16': 'i-0cd794c59c2534969',
    'GrowLiveClient18': 'i-0313545ae9b2ffc8c',
    'GrowLiveCms': 'i-042a7ffab2d9f601c'
    
}

rds_unit_list = ['CPU Utilization (%)', 'Count', 'Count/Second', 'Count/Second']
rds_metric_list = ['CPUUtilization', 'DatabaseConnections', 'ReadIOPS', 'WriteIOPS']
rds_dict = {
    'growlivenew-instance-1': 'growlivenew-instance-1',
    'growlivenew-instance-2': 'growlivenew-instance-2',
    'growlivenew-instance-3': 'growlivenew-instance-3',
    'growlivenew-instance-5': 'growlivenew-instance-5'
}

redis_unit_list = ['CPU Utilization (%)', 'Count', 'Count', 'Count']
redis_metric_list = ['EngineCPUUtilization', 'CurrConnections', 'CurrItems', 'CacheHits']
redis_dict = {
    'growredis-001': 'growredis-001',
    'growredis-002': 'growredis-002'
}

def get_cloudwatch_metric_graphs(namespace, metric_name, identifier, instance_id, unit):
    global cell_tracking
    response = cloudwatch.get_metric_statistics(
        Namespace=namespace,
        MetricName=metric_name,
        Dimensions=[{
            'Name': identifier,
            'Value': instance_id
        }],
        StartTime=week_ago - timedelta(days=7),
        EndTime=week_ago,
        Period=1200,
        Statistics=['Average']
    )
    response['Datapoints'].sort(key=lambda x: x['Timestamp'])
    values = [point['Average'] for point in response['Datapoints']]
    value_averages_last_week = sum(values) / len(values)
    response = cloudwatch.get_metric_statistics(
        Namespace=namespace,
        MetricName=metric_name,
        Dimensions=[{
            'Name': identifier,
            'Value': instance_id
        }],
        StartTime=week_ago,
        EndTime=today,
        Period=1200,
        Statistics=['Average']
    )

    response['Datapoints'].sort(key=lambda x: x['Timestamp'])
    values = [point['Average'] for point in response['Datapoints']]
    value_averages = sum(values) / len(values)
    if metric_name == "NetworkIn" or metric_name == "NetworkOut":
        value_averages = value_averages / 1024 / 1024
        value_averages_last_week = value_averages_last_week / 1024 / 1024

    # Prepare data
    time_stamps = [point['Timestamp'] for point in response['Datapoints']]

    # Create a plot
    plt.figure(figsize=(14, 4))
    plt.plot_date(time_stamps, values, '-')
    print(time_stamps, values)

    # Format the plot
    plt.title('AWS CloudWatch Metric - ' + metric_name)
    plt.xlabel('Time')
    plt.ylabel(unit)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.gca().xaxis.set_major_formatter(DateFormatter('%Y-%m-%d'))
    plt.gcf().autofmt_xdate()  # Rotation

    # Save the plot as an image
    plt.savefig(os.path.join(script_dir, "." + instance_id + metric_name + ".png"))
    plt.close()
    img = Image(os.path.join(script_dir, "." + instance_id + metric_name + ".png"))

    ws.add_image(img, 'A' + str(cell_tracking))  # Adjust the cell location as needed
    ws_pointer = cell_tracking // 20 + 2
    ws["X" + str(ws_pointer)] = metric_name
    ws["Y" + str(ws_pointer)] = value_averages
    ws["Z" + str(ws_pointer)] = value_averages_last_week

    cell_tracking = cell_tracking + 20
def get_metrics(namespace, dict, metric_list, unit_list, indentifier):
    global ws
    global cell_tracking
    for instance in dict.values():
        for key, value in dict.items():
            if value == instance:
                ws = wb.create_sheet(title=key)
                ws["Y1"] = "Last Week"
                ws["Z1"] = "This Week"
        for i, metric in enumerate(metric_list):
            get_cloudwatch_metric_graphs(namespace, metric, indentifier, instance, unit_list[i])
        ws["AA1"] = 'Anomaly'
        ws["AA2"] = '=Z2-Y2'
        ws["AA3"] = '=Z3-Y3'
        ws["AA4"] = '=Z4-Y4'
        ws["AA5"] = '=Z5-Y5'
        cell_tracking = 1

get_metrics('AWS/EC2', ec2_dict, ec2_metric_list, ec2_unit_list, 'InstanceId')
get_metrics('AWS/RDS', rds_dict, rds_metric_list, rds_unit_list, 'DBInstanceIdentifier')
get_metrics('AWS/ElastiCache', redis_dict, redis_metric_list, redis_unit_list, 'CacheClusterId')

today = (datetime.now()).strftime("%d-%m-%Y")
week_ago = week_ago.strftime("%d-%m-%Y")

wb.remove(wb["Sheet"])
wb.save(os.path.join(script_dir, "meshulam_anomaly_" + week_ago + "-" + today + ".xlsx"))

all_items = os.listdir(script_dir)
png_files = [item for item in all_items if item.endswith('.png')]
for file_name in png_files:
    file_path = os.path.join(script_dir, file_name)
    try:
        os.remove(file_path)
        print(f"Deleted: {file_path}")
    except Exception as e:
        print(f"Error deleting {file_path}: {e}")